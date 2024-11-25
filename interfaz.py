# -- coding: utf-8 --
import tkinter as tk
from osgeo import ogr
from tkinter import ttk,filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
import os
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook




class InterfazGrafica:
    def __init__(self, root, app):
        self.root = root
        self.app = app
        self.root.title("Carga Masiva")
        self.root.geometry("800x400")
        self.root.state('zoomed')
        self.gdb_path = tk.StringVar()
        self.excel_file_path = tk.StringVar()
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=1, fill="both")
        
        # Crear las pestañas
        self.tab_validaciones = tk.Frame(self.notebook, bg='#FFFFFF')
        self.tab_consolidar_carpeta = tk.Frame(self.notebook, bg='#FFFFFF')
        self.tab_consolidar_ph_nph = tk.Frame(self.notebook, bg='#FFFFFF')
        self.tab_convertir_gdb = tk.Frame(self.notebook, bg='#FFFFFF')
        
        self.notebook.add(self.tab_validaciones, text="Validaciones")
        self.notebook.add(self.tab_consolidar_carpeta, text="Consolidar carpeta")
        self.notebook.add(self.tab_consolidar_ph_nph, text="Consolidar PH y NPH")
        self.notebook.add(self.tab_convertir_gdb, text="Convertir GDB to PACKAGES")



        #ruta_imagen = os.path.join(os.path.dirname(__file__), "assets", "Logo_Conestudios.png")
        #self.background_image = self.crear_imagen_semitransparente(ruta_imagen, 0.1)
        # Configurar la pestaña de validaciones
        self.configurar_pestania_validaciones()
        self.create_consolidar_carpeta_tab()
        self.create_consolidar_ph_nph_tab()
        self.Configurar_Pestaña_Convertir()
        
    def configurar_pestania_validaciones(self):
    # Cargar la imagen y configurarla como fondo en la pestaña de validaciones
        #self.background_label = tk.Label(self.tab_validaciones, image=self.background_image)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)

        # Crear los widgets dentro de la pestaña "Validaciones"
        frame_nph = tk.Frame(self.tab_validaciones, bg='#7ea7b9')
        frame_nph.pack(pady=20)

        tk.Label(frame_nph, text="Carga Masiva NPH:", font="arial 12 bold", bg='#7ea7b9').pack(side=tk.LEFT, padx=10)
        self.archivo_entry_nph = tk.Entry(frame_nph, width=50)
        self.archivo_entry_nph.pack(side=tk.LEFT, padx=10)
        self.boton_nph = tk.Button(frame_nph, text="Seleccionar Archivo Carga Masiva", command=self.seleccionar_archivo_nph)
        self.boton_nph.pack(side=tk.LEFT, padx=10)
        

        frame_botones = tk.Frame(self.tab_validaciones, bg='#7ea7b9')
        frame_botones.pack(side=tk.BOTTOM, pady=10)

        self.boton_procesar = tk.Button(frame_botones, text="Procesar", font="Arial 16 bold", command=None, state=tk.DISABLED)
        self.boton_procesar.pack(side=tk.LEFT, padx=(0, 20))

        self.boton_limpiar = tk.Button(frame_botones, text="Limpiar", font="Arial 16 bold", command=self.limpiar_seleccion)
        self.boton_limpiar.pack(side=tk.LEFT)

    
    
    def crear_imagen_semitransparente(self, image_path, alpha):
        image = Image.open(image_path)
        image = image.convert("RGBA")
        data = image.getdata()

        new_data = [(item[:-1] + (int(255 * alpha),)) for item in data]
        image.putdata(new_data)
        image = image.resize((self.root.winfo_screenwidth(), self.root.winfo_screenheight()), Image.LANCZOS)
        return ImageTk.PhotoImage(image)    
        
    def create_consolidar_carpeta_tab(self):
        #self.background_label = tk.Label(self.tab_consolidar_carpeta, image=self.background_image)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        # Variables
        self.folder_path = tk.StringVar()
        
        # Instrucciones y entrada de carpeta
        instructions = tk.Label(self.tab_consolidar_carpeta, text="Seleccione la carpeta con los archivos Excel a consolidar")
        instructions.pack(pady=10)
        
        folder_frame = tk.Frame(self.tab_consolidar_carpeta)
        folder_frame.pack(pady=10)
        
        self.folder_entry = tk.Entry(folder_frame, textvariable=self.folder_path, width=50)
        self.folder_entry.pack(side=tk.LEFT, padx=(0, 10))
        browse_btn = tk.Button(folder_frame, text="Buscar", command=self.browse_folder)
        browse_btn.pack(side=tk.LEFT)
        
        # Botón para consolidar
        consolidate_btn = tk.Button(self.tab_consolidar_carpeta, text="Consolidar carpeta", command=self.consolidar_carpeta)
        consolidate_btn.pack(pady=20)
    
    def create_consolidar_ph_nph_tab(self):
        #self.background_label = tk.Label(self.tab_consolidar_ph_nph, image=self.background_image)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        # Variables
        self.ruta_archivo_1 = ""
        self.ruta_archivo_2 = ""
        self.ruta_guardado = ""
        
        # Etiquetas y cuadros de entrada
        tk.Label(self.tab_consolidar_ph_nph, text="Archivo NPH:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.entry_archivo_1 = tk.Entry(self.tab_consolidar_ph_nph, width=50)
        self.entry_archivo_1.grid(row=0, column=1, padx=10, pady=10)
        tk.Button(self.tab_consolidar_ph_nph, text="Seleccionar", command=self.seleccionar_archivo_1).grid(row=0, column=2, padx=10, pady=10)

        tk.Label(self.tab_consolidar_ph_nph, text="Archivo PH:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.entry_archivo_2 = tk.Entry(self.tab_consolidar_ph_nph, width=50)
        self.entry_archivo_2.grid(row=1, column=1, padx=10, pady=10)
        tk.Button(self.tab_consolidar_ph_nph, text="Seleccionar", command=self.seleccionar_archivo_2).grid(row=1, column=2, padx=10, pady=10)

        tk.Label(self.tab_consolidar_ph_nph, text="Guardar como:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        self.entry_guardado = tk.Entry(self.tab_consolidar_ph_nph, width=50)
        self.entry_guardado.grid(row=2, column=1, padx=10, pady=10)
        tk.Button(self.tab_consolidar_ph_nph, text="Seleccionar", command=self.seleccionar_guardado).grid(row=2, column=2, padx=10, pady=10)
        
        # Botón para ejecutar la consolidación
        tk.Button(self.tab_consolidar_ph_nph, text="Consolidar archivos", command=self.consolidar_archivos).grid(row=3, column=1, padx=10, pady=20)
        tk.Button(self.tab_consolidar_ph_nph, text="Generar versión app", command=self.generar_version_app).grid(row=4, column=1, padx=10, pady=20)
        
    def Configurar_Pestaña_Convertir(self):
        #self.background_label = tk.Label(self.tab_convertir_gdb, image=self.background_image)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        #self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        self.gdb_path = tk.StringVar()
        
        tk.Label(self.tab_convertir_gdb, text="Seleccionar",font=("Arial", 12)).pack(pady=10)
        self.entry_convertir = tk.Entry( self.tab_convertir_gdb, width=50)
        self.entry_convertir.pack(pady=5)

        tk.Button(self.tab_convertir_gdb, text="Seleccionar carpeta .gdb", command=self.select_gdb_folder).pack(pady=5)
        tk.Button(self.tab_convertir_gdb, text="Iniciar Conversión", command=self.iniciar_conversion).pack(pady=20)

        
    def select_gdb_folder(self):
        gdb_folder = filedialog.askdirectory(title="Seleccionar carpeta .gdb")
        if gdb_folder:
            self.gdb_path.set(gdb_folder)

    def iniciar_conversion(self):
        if not self.gdb_path.get():
            messagebox.showwarning("Advertencia", "Por favor, selecciona la carpeta GDB.")
            return

        output_folder = filedialog.askdirectory(title="Seleccionar carpeta de destino para los GPKG")

        if not output_folder:
            return

        # Ejecutar la conversión
        self.convert_gdb_to_gpkg(self.gdb_path.get(), output_folder)

    def convert_gdb_to_gpkg(self, gdb_folder, output_folder):
        driver = ogr.GetDriverByName('OpenFileGDB')
        if driver is None:
            messagebox.showerror("Error", "Driver OpenFileGDB no está disponible.")
            return

        # Abre la carpeta GDB como un geodatabase
        gdb = driver.Open(gdb_folder, 0)
        if not gdb:
            messagebox.showerror("Error", f"No se pudo abrir la geodatabase GDB: {gdb_folder}")
            return

        output_driver = ogr.GetDriverByName('GPKG')
        if output_driver is None:
            messagebox.showerror("Error", "No se pudo encontrar el driver de salida.")
            return

        for i in range(gdb.GetLayerCount()):
            layer = gdb.GetLayerByIndex(i)
            layer_name = layer.GetName()

            # Definir la ruta de salida para cada capa con el mismo nombre de la capa
            output_path = os.path.join(output_folder, f"{layer_name}.gpkg")

            # Eliminar el archivo si ya existe
            if os.path.exists(output_path):
                os.remove(output_path)

            # Crear un archivo GPKG por capa
            output_layer = output_driver.CreateDataSource(output_path)
            if output_layer is None:
                messagebox.showerror("Error", f"No se pudo crear el archivo de salida para la capa {layer_name}.")
                continue

            # Copiar la capa al archivo GPKG
            output_layer.CopyLayer(layer, layer_name)

            # Cerrar el DataSource de salida
            output_layer = None

        messagebox.showinfo("Éxito", "Conversión completada para todas las capas.")

        
    
    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path.set(folder_selected)
        
    def format_number(self, value):
        """
        Convierte números grandes a texto para evitar notación científica
        """
        if isinstance(value, (int, float, np.integer, np.floating)):
            if abs(value) >= 1e10:  # Para números mayores o iguales a 10 billones
                return str(int(value))  # Convertir a texto sin decimales
        return value

    def read_excel_preserve_numbers(self, file_path, sheet_name):
        """
        Lee un archivo Excel preservando los números largos como texto
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        data = []
        headers = []
        
        # Leer encabezados
        for cell in sheet[1]:
            headers.append(cell.value)
            
        # Leer datos
        for row in sheet.iter_rows(min_row=2):
            row_data = []
            for cell in row:
                value = cell.value
                # Formatear el valor si es necesario
                formatted_value = self.format_number(value)
                row_data.append(formatted_value)
            data.append(row_data)
            
        # Crear DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df

    def make_unique_columns(self, df):
        """
        Asegura que los nombres de las columnas sean únicos.
        Si hay columnas duplicadas, les añade un sufijo incremental.
        """
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            cols[cols == dup] = [dup + f"_{i}" if i != 0 else dup for i in range(sum(cols == dup))]
        df.columns = cols
        return df
    
    def consolidar_carpeta(self):
        if not self.folder_path.get():
            messagebox.showerror("Error", "Por favor seleccione una carpeta")
            return

        try:
            folder = Path(self.folder_path.get())
            excel_files = list(folder.glob("*.xlsx"))

            if not excel_files:
                messagebox.showerror("Error", "No se encontraron archivos Excel en la carpeta seleccionada")
                return

            first_file = openpyxl.load_workbook(excel_files[0], data_only=True)
            sheet_names = first_file.sheetnames

            consolidated_wb = openpyxl.Workbook()
            consolidated_wb.remove(consolidated_wb.active)

            for sheet_name in sheet_names:
                # Excluir la hoja "Listas" de la consolidación
                if sheet_name == "Listas":
                    # Copiar la hoja "Listas" tal como está desde el primer archivo
                    data = self.read_excel_preserve_numbers(excel_files[0], sheet_name)
                    consolidated_sheet = consolidated_wb.create_sheet(title=sheet_name)
                    consolidated_sheet.append(data.columns.tolist())  # Agregar encabezados

                    # Agregar datos de la hoja original
                    for row in data.itertuples(index=False, name=None):
                        consolidated_sheet.append(row)
                    
                    continue  # Pasar a la siguiente hoja sin consolidar

                if sheet_name == "Leer":
                    # Copiar la hoja "Leer" tal como está desde el primer archivo
                    data = self.read_excel_preserve_numbers(excel_files[0], sheet_name)
                    consolidated_sheet = consolidated_wb.create_sheet(title=sheet_name)
                    consolidated_sheet.append(data.columns.tolist())  # Agregar encabezados

                    # Agregar datos de la hoja original
                    for row in data.itertuples(index=False, name=None):
                        consolidated_sheet.append(row)

                    continue  # Pasar a la siguiente hoja sin consolidar

                # Proceso normal de consolidación para las demás hojas
                consolidated_sheet = consolidated_wb.create_sheet(title=sheet_name)
                dfs = []
                data_validations = []

                for excel_file in excel_files:
                    df = self.read_excel_preserve_numbers(excel_file, sheet_name)

                    # Asegurar que los nombres de las columnas sean únicos
                    df = self.make_unique_columns(df)

                    # Resetear el índice para evitar duplicados
                    df.reset_index(drop=True, inplace=True)

                    # Eliminar filas duplicadas
                    df = df.drop_duplicates().reset_index(drop=True)

                    
                    # Agregar la columna 'Radicado' a todas las hojas
                    df['Radicado'] = os.path.basename(excel_file)

                    # Si la columna 'Npn' existe, crear la columna 'NPN_TERRENO'
                    if 'Npn' in df.columns:
                        df['NPN_TERRENO'] = df['Npn'].astype(str).str[:21]

                    # Agregar el DataFrame procesado a la lista
                    dfs.append(df)
                   

                    # Cargar el libro fuente para copiar validaciones de datos
                    source_wb = openpyxl.load_workbook(excel_file, data_only=False)
                    source_sheet = source_wb[sheet_name]

                    # Copiar validaciones de datos, evitando duplicados
                    for dv in source_sheet.data_validations.dataValidation:
                        if dv not in data_validations:
                            data_validations.append(dv)

                # Concatenar todos los DataFrames asegurando índices únicos
                consolidated_df = pd.concat(dfs, ignore_index=True, sort=False)

                # Aplicar formato de número para evitar notación científica
                for column in consolidated_df.columns:
                    consolidated_df[column] = consolidated_df[column].apply(self.format_number)

                # Guardar en un archivo Excel temporal
                with pd.ExcelWriter("temp_consolidated.xlsx", engine='openpyxl') as writer:
                    consolidated_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Cargar desde el archivo temporal y transferir a la hoja consolidada
                temp_wb = openpyxl.load_workbook("temp_consolidated.xlsx")
                temp_sheet = temp_wb[sheet_name]

                for row in temp_sheet.iter_rows():
                    for cell in row:
                        consolidated_sheet.cell(
                            row=cell.row,
                            column=cell.column,
                            value=cell.value
                        )

                # Aplicar validaciones de datos a la hoja consolidada
                for dv in data_validations:
                    consolidated_dv = DataValidation(
                        type=dv.type,
                        formula1=dv.formula1,
                        formula2=dv.formula2,
                        allow_blank=dv.allow_blank,
                        showDropDown=dv.showDropDown,
                        showErrorMessage=dv.showErrorMessage,
                        errorTitle=dv.errorTitle,
                        error=dv.error,
                        promptTitle=dv.promptTitle,
                        prompt=dv.prompt
                    )
                    consolidated_sheet.add_data_validation(consolidated_dv)
                    for range in dv.ranges:
                        consolidated_dv.add(range)

                # Ajustar el ancho de las columnas basado en la longitud máxima de las celdas
                for column in temp_sheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                    adjusted_width = (max_length + 2)
                    consolidated_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            # Guardar el libro consolidado final
            output_path = folder / "Consolidado.xlsx"
            consolidated_wb.save(output_path)

            # Eliminar el archivo temporal
            if os.path.exists("temp_consolidated.xlsx"):
                os.remove("temp_consolidated.xlsx")

            messagebox.showinfo(
                "Éxito",
                f"Archivos consolidados exitosamente con validaciones.\nArchivo guardado como: {output_path}"
            )

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error durante la consolidación:\n{str(e)}")

        
    def seleccionar_archivo_1(self):
        self.ruta_archivo_1 = filedialog.askopenfilename(title="Selecciona el primer archivo de Excel", filetypes=[("Excel files", "*.xlsx")])
        self.entry_archivo_1.delete(0, tk.END)
        self.entry_archivo_1.insert(0, self.ruta_archivo_1)
    
    def seleccionar_archivo_2(self):
        self.ruta_archivo_2 = filedialog.askopenfilename(title="Selecciona el segundo archivo de Excel", filetypes=[("Excel files", "*.xlsx")])
        self.entry_archivo_2.delete(0, tk.END)
        self.entry_archivo_2.insert(0, self.ruta_archivo_2)

    def seleccionar_guardado(self):
        self.ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Guardar archivo consolidado como", filetypes=[("Excel files", "*.xlsx")])
        self.entry_guardado.delete(0, tk.END)
        self.entry_guardado.insert(0, self.ruta_guardado)
        

    def combinar_hojas(self, lista_hojas, data_archivo_1, data_archivo_2):
        df_consolidado = pd.DataFrame()
        for hoja in lista_hojas:
            df_consolidado = pd.concat([df_consolidado, data_archivo_1.get(hoja, pd.DataFrame()), data_archivo_2.get(hoja, pd.DataFrame())], axis=0, ignore_index=True)
        return df_consolidado

    
    def consolidar_archivos(self):
        if not self.ruta_archivo_1 or not self.ruta_archivo_2 or not self.ruta_guardado:
            messagebox.showwarning("Selección incompleta", "Por favor selecciona los dos archivos de Excel y la ubicación para guardar el archivo consolidado.")
            return

        # Definir las correspondencias entre hojas
        correspondencias = {
            'Fichas': ['Fichas', 'Ficha', 'FichasPrediales'],  
            'Propietarios': ['Propietarios'],
            'Construcciones': ['Construcciones', 'ConstruccionesFicha'],
            'CalificacionesConstrucciones': ['CalificacionesConstrucciones', 'CalificacionesConsFicha'],
            'ConstruccionesGenerales': ['ConstruccionesGenerales', 'ConstruccionGeneralFicha'],
            'Colindantes': ['Colindantes', 'ColindantesFicha'],
            'ZonasHomogeneas': ['ZonasHomogeneas'],
            'CartografiaInformacionGrafica': ['CartografiaInformacionGrafica'],
            'Listas': ['Listas']
        }

        # Leer todas las hojas de ambos archivos
        data_archivo_1 = pd.read_excel(self.ruta_archivo_1, sheet_name=None, dtype=str)
        data_archivo_2 = pd.read_excel(self.ruta_archivo_2, sheet_name=None, dtype=str)

        # Renombrar columnas en el archivo 2 para mantener la consistencia con el archivo 1 MatriculaMatriz CalificacionesConsFicha
        if "Ficha" in data_archivo_2:
            data_archivo_2["Ficha"].rename(columns={'NumCedCatastral': 'NumCedulaCatastral'}, inplace=True)
            data_archivo_2["Ficha"].rename(columns={'MatriculaMatriz': 'MatriculaInmobiliaria'}, inplace=True)
        if "FichasPrediales" in data_archivo_2:
            data_archivo_2["FichasPrediales"].rename(columns={'DestinoEconomico': 'DestinoEcconomico','Puntos ':'Puntos'}, inplace=True)
        if "ConstruccionesFicha" in data_archivo_2:
            data_archivo_2["ConstruccionesFicha"].rename(columns={'Secuencia': 'secuencia', 'IdentificadorUso': 'IdUso', 'PorcentajeConstruccion': 'PorcentajeConstruido','Puntos':'Puntos '}, inplace=True)        
        if "CalificacionesConstrucciones" in data_archivo_2:
            data_archivo_2["CalificacionesConstrucciones"].rename(columns={'CubrimientoMuro': 'Cubrimiento Muro'}, inplace=True)
            data_archivo_2["CalificacionesConstrucciones"].rename(columns={'CubrimientoMuro': 'Cubrimiento Muro'}, inplace=True)

        # Diccionario para almacenar los dataframes consolidados
        consolidado = {}

        # Consolidar las hojas correspondientes
        for hoja_destino, hojas_fuente in correspondencias.items():
            consolidado[hoja_destino] = self.combinar_hojas(hojas_fuente, data_archivo_1, data_archivo_2)

        # Agregar funcionalidad para la hoja Construcciones
        if 'Construcciones' in consolidado:
            construcciones_df = consolidado['Construcciones']

            # Agregar la columna FHNC
            construcciones_df['FHNC'] = construcciones_df['NroFicha'] + "-" + construcciones_df['NumeroConstruccion']

            # Buscar en Fichas para obtener el valor de NpnConst
            if 'Fichas' in consolidado:
                fichas_df = consolidado['Fichas']
                construcciones_df = construcciones_df.merge(
                    fichas_df[['NroFicha', 'Npn']],
                    on='NroFicha',
                    how='left',
                    suffixes=('', '_fichas')
                )
                construcciones_df['NpnConst'] = construcciones_df['Npn'].astype(str).str[:21] + "00000000-" + construcciones_df['NumeroConstruccion']
                construcciones_df.drop(columns=['Npn'], inplace=True)  # Eliminar la columna Npn tras usarla
                consolidado['Construcciones'] = construcciones_df

            # Buscar en Fichas para obtener el valor de MatriculaInmobiliaria
            '''if 'Fichas' in consolidado:
                fichas_df = consolidado['Fichas']
                construcciones_df = construcciones_df.merge(
                    fichas_df[['NroFicha', 'MatriculaInmobiliaria']],
                    on='NroFicha',
                    how='left',
                    suffixes=('', '_fichas')
                )
                construcciones_df['MatriculaInmobiliaria'] = construcciones_df['MatriculaInmobiliaria']
                consolidado['Propietarios'] = construcciones_df

            # Buscar en Fichas para obtener el valor de tomo
            if 'Fichas' in consolidado:
                fichas_df = consolidado['Fichas']
                construcciones_df = construcciones_df.merge(
                    fichas_df[['NroFicha', 'Tomo']],
                    on='NroFicha',
                    how='left',
                    suffixes=('', '_fichas')
                )
                construcciones_df['Tomo'] = construcciones_df['Tomo']
                consolidado['Propietarios'] = construcciones_df'''

        if 'CalificacionesConstrucciones' in consolidado and 'Construcciones' in consolidado:
            calificaciones_df = consolidado['CalificacionesConstrucciones']  
            calificaciones_df.rename(columns={'Secuencia': 'secuencia'}, inplace=True)

            construcciones_df = consolidado['Construcciones']

            # Hacer un merge usando la columna 'secuencia'
            calificaciones_df = calificaciones_df.merge(
                construcciones_df[['secuencia', 'FHNC', 'NpnConst','TipoConstruccion']],
                on='secuencia',
                how='left'
            )

            consolidado['CalificacionesConstrucciones'] = calificaciones_df

        if 'Fichas' in consolidado:
            fichas_df = consolidado['Fichas']

            
            # Llenar el campo Corregimiento si está vacío
            fichas_df['Corregimiento'] = fichas_df['Corregimiento'].where(
                ~fichas_df['Corregimiento'].isna(),  # Mantener valores que NO están vacíos
                fichas_df['NumCedulaCatastral'].str[4:7]  # Extraer caracteres 5 al 7
            )

            # Llenar el campo Barrio si está vacío
            fichas_df['Barrio'] = fichas_df['Barrio'].where(
                ~fichas_df['Barrio'].isna(),
                fichas_df['NumCedulaCatastral'].str[7:10]  # Extraer caracteres 8 al 10
            )

            # Llenar el campo Manzana si está vacío
            fichas_df['ManzanVereda'] = fichas_df['ManzanVereda'].where(
                ~fichas_df['ManzanVereda'].isna(),
                fichas_df['NumCedulaCatastral'].str[10:14]  # Extraer caracteres 11 al 14
            )

            # Llenar el campo Predio si está vacío
            fichas_df['Predio'] = fichas_df['Predio'].where(
                ~fichas_df['Predio'].isna(),
                fichas_df['NumCedulaCatastral'].str[14:19]  # Extraer desde el carácter 15 en adelante
            )
        

        if 'Fichas' in consolidado and 'Propietarios' in consolidado:
            fichas_df = consolidado['Fichas']
            propietarios_df = consolidado['Propietarios']

            # Realizar merge entre Fichas y Propietarios usando NroFicha
            propietarios_df = propietarios_df.merge(
                fichas_df[['NroFicha', 'MatriculaInmobiliaria','Tomo']],
                on='NroFicha',
                how='left'
            )

            # Actualizar el DataFrame de Propietarios con la columna MatriculaInmobiliaria
            consolidado['Propietarios'] = propietarios_df
            print("Campo MatriculaInmobiliaria copiado de Fichas a Propietarios según NroFicha.")
          
        

        # Actualizar el consolidado con la nueva versión de fichas_df
        consolidado['Fichas'] = fichas_df

        # Agregar las hojas que no tienen correspondencia
        hojas_no_correspondidas = set(data_archivo_2.keys()).difference([item for sublist in correspondencias.values() for item in sublist])
        for hoja in hojas_no_correspondidas:
            consolidado[hoja] = data_archivo_2[hoja]

        # Guardar el archivo consolidado
        ruta_consolidada = self.ruta_guardado
        with pd.ExcelWriter(ruta_consolidada, engine='xlsxwriter') as writer:
            for nombre_hoja, df in consolidado.items():
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        # Cargar el archivo consolidado con openpyxl para agregar las listas desplegables
        try:
            wb_consolidado = load_workbook(ruta_consolidada)

            # Copiar validaciones de cada archivo original
            for ruta_archivo in [self.ruta_archivo_1, self.ruta_archivo_2]:
                wb_original = load_workbook(ruta_archivo)

                for nombre_hoja in wb_original.sheetnames:
                    hoja_original = wb_original[nombre_hoja]

                    # Verifica si la hoja existe en el archivo consolidado
                    if nombre_hoja in wb_consolidado.sheetnames:
                        hoja_consolidada = wb_consolidado[nombre_hoja]

                        # Copiar las validaciones de datos
                        if hoja_original.data_validations:
                            for dv in hoja_original.data_validations.dataValidation:
                                new_dv = DataValidation(
                                    type=dv.type, formula1=dv.formula1, formula2=dv.formula2,
                                    showDropDown=dv.showDropDown, allowBlank=dv.allowBlank
                                )
                                new_dv.sqref = dv.sqref
                                hoja_consolidada.add_data_validation(new_dv)

            # Guardar el archivo con las validaciones copiadas
            wb_consolidado.save(ruta_consolidada)

        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar el archivo: {e}")
            return

        messagebox.showinfo("Consolidación completada", f"El archivo consolidado se ha guardado en {ruta_consolidada}")
    def generar_version_app(self):
        if not self.ruta_guardado:
            messagebox.showwarning("Selección incompleta", "Por favor selecciona una ubicación para guardar el archivo consolidado.")
            return

        try:
            ruta_consolidada = self.ruta_guardado
            output_path = ruta_consolidada.replace(".xlsx", "_version_app.xlsx")

            # Lista de columnas a omitir según libro y hoja
            omisiones = {
                'Fichas': ['TipoMutacion', 'DireccionReferencia','ObjectId','ObjectIdEdificio','NumeroRadicado','Resolucion','FechaResolucion','Lado'],
                'Propietarios': ['CodigoFideicomiso','SiglaComercial','DepartamentoPersona','MunicipioPersona','Direccion'],
                'Construcciones': ['ObjectIdConstruccion','ObjectIdModelo',],
                # Agrega más hojas y columnas según sea necesario
            }

            # Cargar el archivo consolidado
            wb = load_workbook(ruta_consolidada)
            for sheet_name, columns_to_omit in omisiones.items():
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # Obtener índices de columnas a eliminar
                    column_indices = [
                        idx for idx, cell in enumerate(sheet[1], start=1) if cell.value in columns_to_omit
                    ]
                    # Eliminar columnas desde la última hacia la primera
                    for col_idx in sorted(column_indices, reverse=True):
                        sheet.delete_cols(col_idx)

            # Guardar el archivo modificado
            wb.save(output_path)
            messagebox.showinfo("Generación completada", f"El archivo generado se ha guardado en {output_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error durante la generación del archivo:\n{str(e)}")
        

    def seleccionar_archivo_nph(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.archivo_entry_nph.delete(0, tk.END)
            self.archivo_entry_nph.insert(0, filename)
            self.boton_procesar.config(command=self.app.procesar_archivo, state=tk.NORMAL)
            self.boton_rph.config(state=tk.DISABLED)
        else:
            self.boton_rph.config(state=tk.NORMAL)
                
           
    def select_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.excel_file_path.set(filename)
    
    def limpiar_seleccion(self):
        self.archivo_entry_nph.delete(0, tk.END)
        self.archivo_entry_rph.delete(0, tk.END)
        self.boton_rph.config(state=tk.NORMAL)
        self.boton_nph.config(state=tk.NORMAL)
    
    def limpiar_seleccion_fichas(self):
        self.archivo_entry_fichas.delete(0, tk.END)
        self.boton_procesar_fichas.config(state=tk.DISABLED)
    
    def select_gdb(self):
        gdb_folder = filedialog.askdirectory(title="Seleccionar carpeta .gdb")
        if gdb_folder:
            self.gdb_path.set(gdb_folder)
            

    

    def crear_imagen_semitransparente(self, image_path, alpha):
        image = Image.open(image_path)
        image = image.convert("RGBA")
        data = image.getdata()

        new_data = [(item[:-1] + (int(255 * alpha),)) for item in data]
        image.putdata(new_data)
        image = image.resize((self.root.winfo_screenwidth(), self.root.winfo_screenheight()), Image.LANCZOS)
        return ImageTk.PhotoImage(image)
    
    

    